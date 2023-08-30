import openpyxl

print('Opening Workbook...')
wb = openpyxl.load_workbook('IA_Marks.xlsx', data_only=True)
sheet = wb['Sheet1']
print('Rows in the workbook:', sheet.max_row)
print('Reading rows...')
for row in range(1, sheet.max_row+1):
    USN = sheet['A' + str(row)].value
    Name = sheet['B' + str(row)].value
    test1 = sheet['C' + str(row)].value
    test2 = sheet['D' + str(row)].value
    test3 = sheet['E' + str(row)].value
    tot = sheet['F' + str(row)].value
    avg = sheet['G' + str(row)].value
    print("%4s %12s %20s %7s %7s %7s %7s %7.4s" % (row, USN, Name, test1, test2, test3, tot, avg))


row += 1
sheet['A' + str(row)] = input("Enter USN: ")
sheet['B' + str(row)] = input("Enter Name: ")
test1 = sheet['C' + str(row)] = int(input("Test 1 marks: "))
test2 = sheet['D' + str(row)] = int(input("Test 2 marks: "))
test3 = sheet['E' + str(row)] = int(input("Test 3 marks: "))
tot = sheet['F' + str(row)] = test1 + test2 + test3
sheet['G' + str(row)] = tot / 3
wb.save('IA_Marks.xlsx')
wb.close()

print('Opening workbook...')
wb = openpyxl.load_workbook('IA_Marks.xlsx', data_only=True)
sheet = wb['Sheet1']
print('Reading rows...')
for row in range(1, sheet.max_row+1):
    USN = sheet['A' + str(row)].value
    Name = sheet['B' + str(row)].value
    test1 = sheet['C' + str(row)].value
    test2 = sheet['D' + str(row)].value
    test3 = sheet['E' + str(row)].value
    tot = sheet['F' + str(row)].value
    avg = sheet['G' + str(row)].value
    print("%4s %12s %20s %7s %7s %7s %7s %7.2s" % (row, USN, Name, test1, test2, test3, tot, avg))
wb.close()